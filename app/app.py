import tkinter as tk
from tkinter import filedialog, messagebox
import configparser
from bs4 import BeautifulSoup
import requests
from requests.exceptions import RequestException
from .config import get_config_value, get_pdf_path, set_config_value
from .data_list import sites
import app.utils_scrapping as scrapping
from .utils_pdf import download_pdf, delete_pdfs
import datetime
from datetime import timedelta
from datetime import date
from openpyxl import load_workbook, Workbook
import sys
import os
import subprocess
from io import StringIO
from ressources.colors import bg_color, bg_color_light, bg_color, text_light, text_medium, text_dark
import tkinter.messagebox as messagebox
from app.utils_format import check_and_return_value
import threading
import ssl
import locale
from dateutil.easter import easter

from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QVBoxLayout, QWidget, QPushButton, QFileDialog, QTextEdit, QLineEdit
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QInputDialog, QProgressBar
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QTimer
from PyQt5 import QtWidgets, QtCore

config = configparser.ConfigParser()
config.read('config.ini')
locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
now = datetime.datetime.now().date()
yesterday = now - timedelta(days=1)

# Ajd 'vendredi'
day_of_week = now.strftime("%A")

# Hier '01/06/2023
date_yesterday = yesterday.strftime("%d/%m/%Y")

# Hier 'jeudi'
yesterday_day_of_week = yesterday.strftime("%A")

# Hier 'jeudi 01 juin 2023
yesterday_holiday = yesterday.strftime("%A %d %B")

def get_uk_holidays(year):
    # Jours fériés fixes
    holidays_uk = [
        date(year, 1, 1),   # Jour de l'an
        date(year, 12, 25), # Noël
        date(year, 12, 26), # Lendemain de Noël
    ]
    
    # Premier lundi de mai
    may_day = date(year, 5, 1)
    while may_day.weekday() != 0:
        may_day += timedelta(days=1)
    holidays_uk.append(may_day)
    
    # Dernier lundi de mai
    spring_bank_holiday = date(year, 5, 31)
    while spring_bank_holiday.weekday() != 0:
        spring_bank_holiday -= timedelta(days=1)
    holidays_uk.append(spring_bank_holiday)
    
    # Dernier lundi d'août
    summer_bank_holiday = date(year, 8, 31)
    while summer_bank_holiday.weekday() != 0:
        summer_bank_holiday -= timedelta(days=1)
    holidays_uk.append(summer_bank_holiday)
    
    # Jours fériés variables basés sur Pâques
    good_friday = easter(year) - timedelta(days=2)
    holidays_uk.append(good_friday)
    
    easter_monday = easter(year) + timedelta(days=1)
    holidays_uk.append(easter_monday)

    holidays_uk_formatted = [
        holiday.strftime('%A %d %B').lower() for holiday in holidays_uk
    ]
    
    return holidays_uk_formatted


def get_french_holidays(year):
    # Jours fériés fixes
    holidays_french = [
        date(year, 1, 1),   # Jour de l'an
        date(year, 5, 1),   # Fête du travail
        date(year, 5, 8),   # Victoire des alliés
        date(year, 7, 14),  # Fête nationale
        date(year, 8, 15),  # Assomption
        date(year, 11, 1),  # Toussaint
        date(year, 11, 11), # Armistice
        date(year, 12, 25),# Noël
    ]
    
    # Jours fériés variables
    lundi_paques = easter(year) + timedelta(days=1)
    holidays_french.append(lundi_paques)
    
    ascension = easter(year) + timedelta(days=39)
    holidays_french.append(ascension)
    
    pentecote = easter(year) + timedelta(days=50)
    holidays_french.append(pentecote)

    # vendredi_saint = easter(year) - timedelta(days=2)
    # holidays_french.append(vendredi_saint)

    holidays_french_formatted = [
        holiday.strftime('%A %d %B').lower() for holiday in holidays_french
    ]

    return holidays_french_formatted

# Lire le chemin du fichier à partir du fichier config.ini
config = configparser.ConfigParser()
if os.path.exists('config.ini'):
    config.read('config.ini')
    default_path_excel = config.get('SETTINGS', 'excel_path', fallback="")
    default_path_pdf = config.get('SETTINGS', 'pdf_path', fallback="")
    default_path_pdf_name = config.get('SETTINGS', 'name_pdf', fallback="")
else:
    default_path_excel = ""
    default_path_pdf = ""
    default_path_pdf_name = ""

# Ajout de la fenêtre de chargement
class LoadingWindow(QMainWindow):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Chargement...")
        layout = QVBoxLayout()
        self.label = QLabel("Le script s'exécute, veuillez patienter...")
        layout.addWidget(self.label)
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def close(self):
        self.destroy()

class ScrappingSelectionDialog(QtWidgets.QDialog):
    def __init__(self, scrapping_functions, selected_functions=[], parent=None):
        super().__init__(parent)
        self.setWindowTitle('Sélectionner les fonctions de scrapping')
        
        self.layout = QtWidgets.QVBoxLayout(self)
        
        self.checkboxes = []  # Liste pour stocker les checkboxes
        
        for func in scrapping_functions:
            checkbox = QtWidgets.QCheckBox(func)  # Utilisez le nom de la fonction comme label
            if func in selected_functions:
                checkbox.setChecked(True)
            self.layout.addWidget(checkbox)
            self.checkboxes.append(checkbox)  # Ajoutez chaque checkbox à la liste
            
        self.ok_button = QtWidgets.QPushButton('OK', self)
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

    def get_selected_functions(self):
        selected_functions = []
        for checkbox in self.checkboxes:
            if checkbox.isChecked():
                selected_functions.append(checkbox.text())
        return selected_functions

class MyApp(QtWidgets.QWidget):

    def open_scrapping_selection_dialog(self):
        scrapping_functions = [site['func'] for site in sites]  # Créer une liste des noms de fonctions
        dialog = ScrappingSelectionDialog(scrapping_functions, self.selected_scrapping_functions, self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.selected_scrapping_functions = dialog.get_selected_functions()
            self.save_selected_functions()

    def save_selected_functions(self):
        selected_functions_str = ",".join(self.selected_scrapping_functions)
        self.config.set('SETTINGS', 'selected_functions', selected_functions_str)
        with open('config.ini', 'w') as configfile:
            self.config.write(configfile)

    def load_selected_functions(self):
        selected_functions_str = self.config.get('SETTINGS', 'selected_functions', fallback="")
        self.selected_scrapping_functions = selected_functions_str.split(",") if selected_functions_str else []


class MyApp(QtWidgets.QWidget):

    def open_scrapping_selection_dialog(self):
        scrapping_functions = [site['func'] for site in sites]  # Créer une liste des noms de fonctions
        dialog = ScrappingSelectionDialog(scrapping_functions, self.selected_scrapping_functions, self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.selected_scrapping_functions = dialog.get_selected_functions()
            self.save_selected_functions()

    def save_selected_functions(self):
        selected_functions_str = ",".join(self.selected_scrapping_functions)
        self.config.set('SETTINGS', 'selected_functions', selected_functions_str)
        with open('config.ini', 'w') as configfile:
            self.config.write(configfile)

    def load_selected_functions(self):
        selected_functions_str = self.config.get('SETTINGS', 'selected_functions', fallback="")
        self.selected_scrapping_functions = selected_functions_str.split(",") if selected_functions_str else []

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Cours des métaux")
        self.setGeometry(100, 100, 900, 820)
        self.initUI()

        self.layout = QtWidgets.QVBoxLayout(self)

        self.select_scrapping_functions_button = QtWidgets.QPushButton('Sélectionner les fonctions de scrapping', self)
        self.select_scrapping_functions_button.clicked.connect(self.open_scrapping_selection_dialog)
        self.layout.addWidget(self.select_scrapping_functions_button)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')
        self.load_selected_functions()
        auto_start = self.config.getboolean('SETTINGS', 'auto_start', fallback=False)
        self.param1_checkbox.setChecked(auto_start)

        if auto_start:
            if not self.path_pdf.text():
                QMessageBox.warning(
                self, "Chemin PDF manquant",
                "Le chemin d'accès au PDF est manquant. Veuillez le configurer avant de lancer le script.")
            else:    
                self.lancer_script(sites)
                QtCore.QTimer.singleShot(120000, self.close)


    def initUI(self):

      ############ INTERFACE #############

        main_layout = QtWidgets.QHBoxLayout()
        # Layout gauche
        layout = QtWidgets.QVBoxLayout()

        

        # Label pour le chemin d'accès Excel
        self.label_excel_path = QLabel("Chemin d'accès Excel :")
        layout.addWidget(self.label_excel_path)

        # Chemin d'accès excel
        self.path_excel = QtWidgets.QLineEdit(get_config_value("SETTINGS", "excel_path"))  
        self.path_excel.setReadOnly(True)
        layout.addWidget(self.path_excel)

        # Boutons Modifier et Ouvrir excel
        button_layout_excel = QtWidgets.QHBoxLayout()
        self.modify_button_excel = QtWidgets.QPushButton('Modifier')
        self.open_button_excel = QtWidgets.QPushButton('Ouvrir')
        button_layout_excel.addWidget(self.modify_button_excel)
        button_layout_excel.addWidget(self.open_button_excel)
        self.modify_button_excel.setToolTip('Cliquez ici pour modifier le chemin')

        layout.addLayout(button_layout_excel)

        # Label pour le chemin d'accès Excel
        self.label_pdf_path = QLabel("Chemin d'accès PDF :")
        layout.addWidget(self.label_pdf_path)

        # Chemin d'accès PDF
        path_pdf = get_config_value("SETTINGS", "pdf_path")
        self.path_pdf = QtWidgets.QLineEdit(path_pdf)
        self.path_pdf.setReadOnly(True)
        layout.addWidget(self.path_pdf)

        # Boutons Modifier PDF
        button_layout_pdf = QtWidgets.QHBoxLayout()
        self.modify_button_pdf = QtWidgets.QPushButton('Modifier')
        button_layout_pdf.addWidget(self.modify_button_pdf)
        layout.addLayout(button_layout_pdf)

         # Label pour le chemin d'accès Excel
        self.label_name_pdf_path = QLabel("Nom du PDF :")
        layout.addWidget(self.label_name_pdf_path)

        # Nom PDF
        name_pdf = get_config_value("SETTINGS", "name_pdf")
        self.path_namepdf = QtWidgets.QLineEdit(name_pdf)
        self.path_namepdf.setReadOnly(True)
        layout.addWidget(self.path_namepdf)

        # Boutons Modifier nom PDF
        button_layout_namepdf = QtWidgets.QHBoxLayout()
        self.modify_button_namepdf = QtWidgets.QPushButton('Modifier')
        button_layout_namepdf.addWidget(self.modify_button_namepdf)
        layout.addLayout(button_layout_namepdf)

        # Connexion Buttons avec fonctions
        self.modify_button_excel.clicked.connect(self.modify_path_excel)
        self.open_button_excel.clicked.connect(self.open_file_excel)
        self.modify_button_pdf.clicked.connect(self.modify_path_pdf)
        self.modify_button_namepdf.clicked.connect(self.modify_name)

        self.progressbar = QProgressBar(self)
        layout.addWidget(self.progressbar)

        # Création de la section Paramètres
        self.settings_group = QtWidgets.QGroupBox("Paramètres")
        settings_layout = QtWidgets.QVBoxLayout()


        # Bouton pour ouvrir la sélection des fonctions de scrapping
        self.select_scrapping_functions_button = QtWidgets.QPushButton('Sélectionner les Rates à récupérer.')
        self.select_scrapping_functions_button.clicked.connect(self.open_scrapping_selection_dialog)
        settings_layout.addWidget(self.select_scrapping_functions_button)

        # Date
        self.start_date_edit = QtWidgets.QDateEdit(datetime.datetime.now().date())
        self.end_date_edit = QtWidgets.QDateEdit(datetime.datetime.now().date())

        # Champs choix des dates
        self.start_date_edit.setCalendarPopup(True)
        self.end_date_edit.setCalendarPopup(True)

        settings_layout.addWidget(QtWidgets.QLabel("Date de début :"))
        settings_layout.addWidget(self.start_date_edit)
        settings_layout.addWidget(QtWidgets.QLabel("Date de fin :"))
        settings_layout.addWidget(self.end_date_edit)

        # Créez l'objet QCheckBox avant de l'ajouter au layout
        self.use_date_range_checkbox = QtWidgets.QCheckBox("Utiliser la plage de dates")
        self.use_date_range_checkbox.stateChanged.connect(self.toggle_date_widgets)
        settings_layout.addWidget(self.use_date_range_checkbox)

        self.start_date_edit.setEnabled(self.use_date_range_checkbox.isChecked())
        self.end_date_edit.setEnabled(self.use_date_range_checkbox.isChecked())
        

        # Ajout de différents widgets pour les paramètres
        self.param1_checkbox = QtWidgets.QCheckBox("Lancer le script automatiquement au démarrage de l'application.")
        self.param1_checkbox.stateChanged.connect(self.saveSettings)

        # Ajout des widgets au layout des paramètres
        settings_layout.addWidget(self.param1_checkbox)

        # Définition du layout des paramètres comme layout du QGroupBox
        self.settings_group.setLayout(settings_layout)

        # Ajout du QGroupBox au layout principal
        layout.addWidget(self.settings_group)


        main_layout.addLayout(layout)

        # Layout droit (Log)
        log_layout = QtWidgets.QVBoxLayout()
        self.logger = QtWidgets.QTextEdit()
        self.logger.setReadOnly(True)
        log_layout.addWidget(self.logger)

        # Ajout du layout droit au layout principal
        main_layout.addLayout(log_layout)

        # Bouton Lancer
        self.run_button = QtWidgets.QPushButton('Lancer')
        layout.addWidget(self.run_button)
        self.run_button.clicked.connect(lambda: self.lancer_script(sites))

        # Définition du layout principal comme layout de la fenêtre
        self.setLayout(main_layout)

        # INIT #
        self.show()

        self.update_run_button_status(day_of_week)
    
    #############  FONCTIONS ###############

    def toggle_date_widgets(self):
        if self.use_date_range_checkbox.isChecked():
            self.start_date_edit.setEnabled(True)
            self.end_date_edit.setEnabled(True)
        else:
            self.start_date_edit.setEnabled(False)
            self.end_date_edit.setEnabled(False)

    def update_run_button_status(self, day):
        
        if day in ["samedi", "dimanche"]:
            self.run_button.setEnabled(False)
            QMessageBox.information(self, "Jour fermé.", "Jour fermé, le script ne peut être lancé.")
        elif not self.path_pdf.text().strip():
            self.run_button.setEnabled(False)
            QMessageBox.information(self, "Chemin d'accès PDF manquant.", "Veuillez renseigner un chemin d'accès PDF valide.")
        else:
            self.run_button.setEnabled(True)

    def saveSettings(self):
        self.config.read('config.ini')
    
        # Mettez à jour seulement la clé spécifique
        if not self.config.has_section('SETTINGS'):
            self.config.add_section('SETTINGS')
        self.config.set('SETTINGS', 'auto_start', str(self.param1_checkbox.isChecked()))
        
        # Écrivez le fichier de configuration mis à jour
        with open('config.ini', 'w') as configfile:
            self.config.write(configfile)

    def modify_path_excel(self):
        file_dialog = QFileDialog()
        path = file_dialog.getOpenFileName(self, 'Sélectionner un fichier Excel', '', 'Excel Files (*.xlsx *.xls)')[0]
        if path:
            self.path_excel.setText(path)
            # Lire le fichier de configuration existant
            config.read('config.ini')
            # Mettre à jour seulement la clé spécifique
            if not config.has_section('SETTINGS'):
                config.add_section('SETTINGS')
            config.set('SETTINGS', 'excel_path', path)
            # Écrire le fichier de configuration mis à jour
            with open('config.ini', 'w') as configfile:
                config.write(configfile)
            self.log('Chemin modifié.')
    
    def modify_path_pdf(self):
        file_dialog = QFileDialog()
        path = file_dialog.getExistingDirectory(self, 'Sélectionner un dossier')
        old_path = self.path_pdf.text()
        if path:
            self.path_pdf.setText(path)
            # Lire le fichier de configuration existant
            config.read('config.ini')
            # Mettre à jour seulement la clé spécifique
            if not config.has_section('SETTINGS'):
                config.add_section('SETTINGS')
            config.set('SETTINGS', 'pdf_path', path)
            # Écrire le fichier de configuration mis à jour
            with open('config.ini', 'w') as configfile:
                config.write(configfile)
            self.log('Chemin modifié.')
            if not self.restart_app():
                self.path_pdf.setText(old_path)
        self.update_run_button_status(day_of_week)
        

    def modify_name(self):
        new_name, ok = QInputDialog.getText(self, 'Modifier le nom', 'Entrez le nouveau nom:')
        old_name = self.label_name_pdf_path.text()
        if ok and new_name:
            self.path_namepdf.setText(new_name)
            # Lire le fichier de configuration existant
            config.read('config.ini')
            # Mettre à jour seulement la clé spécifique
            if not config.has_section('SETTINGS'):
                config.add_section('SETTINGS')
            config.set('SETTINGS', 'name_pdf', new_name)
            # Écrire le fichier de configuration mis à jour
            with open('config.ini', 'w') as configfile:
                config.write(configfile)
            self.log('Nom modifié.')
            if not self.restart_app():
                self.label_name_pdf_path.setText(old_name)

    def open_file_excel(self):
        # Fonction pour ouvrir le fichier
        try:
            subprocess.run(["start", default_path_excel], shell=True, check=True)
            self.log('Fichier ouvert.')
        except subprocess.CalledProcessError as e:
            self.log('Fichier non trouvé.')

    def restart_app(self):
        """Redémarre l'application."""
        reply = QMessageBox.question(
            self, "Redémarrage requis",
            "L'application doit être redémarrée pour appliquer les changements. Voulez-vous redémarrer maintenant?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Redémarrez l'application
            QApplication.quit()
            os.execl(sys.executable, sys.executable, *sys.argv)
        else:
            return False

    def lancer_script(self, sites):

        replaced_values = {}
        replaced_value_count = 0

        self.progressbar.setMaximum(len(sites))
        self.progressbar.setValue(0)
        holidays_french = get_french_holidays(yesterday.year)
        holidays_uk = get_uk_holidays(yesterday.year)

        # Récupérer les dates de début et de fin de l'interface utilisateur
        if self.use_date_range_checkbox.isChecked():
            start_date = self.start_date_edit.date().toPyDate()
            end_date = self.end_date_edit.date().toPyDate()
        else:
            start_date = end_date = datetime.date.today()

         # Création fichier excel s'il n'existe pas
        excel_path = default_path_excel
        if not excel_path or not os.path.exists(excel_path):
            excel_path = os.path.join(os.getcwd(), "metals_prices.xlsx")
            wb = Workbook()
            for site in sites:
                wb.create_sheet(site['name'])

            wb.save(excel_path)
            set_config_value("SETTINGS", "excel_path", excel_path)
            self.path_excel.setText(excel_path)
        # On charge le fichier s'il existe  
        else:
            wb = load_workbook(excel_path)

        rpa_sheet = wb['RPA'] if 'RPA' in wb.sheetnames else wb.create_sheet('RPA')
        # Clear existing data in "RPA" sheet
        if rpa_sheet.max_row > 1:
            rpa_sheet.delete_rows(2, rpa_sheet.max_row-1)



        txterr = ""
        for site in sites:
            data_extraction_function_name = site['func']
            if data_extraction_function_name not in self.selected_scrapping_functions:
                continue
            try:
                response = requests.get(site['url'], verify=False)
                response.raise_for_status()
            except RequestException as e:
                txterr = f"Erreur de connexion pour le site de {site['name']} : {e}"
                self.log(txterr)
                continue

            soup = BeautifulSoup(response.content, "html.parser")
            if hasattr(scrapping, data_extraction_function_name):
                if site['src'] == 'pdf':
                    download_pdf(response, site['name_pdf'], default_path_pdf)
                else:
                    print('')

                data_extraction_function = getattr(scrapping, data_extraction_function_name)
                sheet = wb[site["name"]]

                extracted_data = data_extraction_function(soup, checkbox_state=self.use_date_range_checkbox.isChecked(), start_date=start_date, end_date=end_date)
                data = None
                # Barre de progression
                self.progressbar.setValue(self.progressbar.value() + 1)
                
                # Si la checkbox est cochée, traiter chaque paire de données extraites
                if self.use_date_range_checkbox.isChecked():
                    for date_day, data in extracted_data:
                        row_number = sheet.max_row + 1
                        sheet.cell(row=row_number, column=1, value=date_day)
                        sheet.cell(row=row_number, column=2, value=data)
                        sheet.cell(row=row_number, column=3, value=site['devise'])
                        sheet.cell(row=row_number, column=4, value=site['unit'])
                else:
                    # Si la checkbox n'est pas cochée, prenez la première paire de données extraites
                    date_day, data = extracted_data
                    row_number = sheet.max_row + 1
                    sheet.cell(row=row_number, column=1, value=date_day)
                    
                    # Vérifier les jours fériés et écrire la valeur appropriée
                    if site['cal'] == 'fr' and date_day not in holidays_french:
                        sheet.cell(row=row_number, column=2, value=data)
                    elif site['cal'] == 'uk' and date_day not in holidays_uk:
                        sheet.cell(row=row_number, column=2, value=data)
                    else:
                        sheet.cell(row=row_number, column=2, value="Jour non valeur")

                    sheet.cell(row=row_number, column=3, value=site['devise'])
                    sheet.cell(row=row_number, column=4, value=site['unit'])

                 # Écrire toutes les valeurs dans l'onglet RPA
                # for date_day, data in extracted_data:
                #     rpa_row_number = rpa_sheet.max_row + 1
                #     rpa_sheet.cell(row=rpa_row_number, column=1, value=site['metal'])
                #     rpa_sheet.cell(row=rpa_row_number, column=2, value=site['name'])
                #     if site['cal'] == 'fr' and not yesterday_holiday in holidays_french:
                #         rpa_sheet.cell(row=rpa_row_number, column=3, value=data)
                #     elif site['cal'] == 'uk' and not yesterday_holiday in holidays_uk:
                #         rpa_sheet.cell(row=rpa_row_number, column=3, value=data)
                #     else:
                #         rpa_sheet.cell(row=rpa_row_number, column=3, value="Jour non valeur")
                #     rpa_sheet.cell(row=rpa_row_number, column=4, value=site['devise'])
                #     rpa_sheet.cell(row=rpa_row_number, column=5, value=site['unit'])

                self.log(txterr)
                wb.save(excel_path)
                print("Saved")
            else:
                print(f'Aucune fonction d\'extraction de données trouvées')
        replaced_message = f"{replaced_value_count} valeurs remplacées : {', '.join(f'{k}: {v}' for k, v in replaced_values.items())}"
        self.log("Script terminé.")
        QMessageBox.information(self, "Information", f"Le script a terminé l'extraction des données et la mise à jour du fichier Excel.\n{replaced_message}")
    
    def log(self, message):
        # Fonction pour log les messages
        self.logger.append(message)